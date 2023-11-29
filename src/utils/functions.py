import time
import string
import random
import hashlib
import openpyxl
import pandas as pd

def generate_unique_token(
        lst_existing_token: list,
        length: int = 32
    ) -> str:
    """
    Metodo empleado para la generacion de Tokens de manera automatica. 
    De manera interna este metodo se encarga de verificar si el token 
    que se esta generando no haya sido generado con anterioridad y se 
    encuentre en uso.

    Args:
    -----

    lst_existing_token: list
        Lista de tokens existentes con la que se comprobara la posible 
        existencia del token que se esta generando en esta funcion.

    length: int
        Default 32
        Este valor puede variar pero se fija alto para evitar la posibilidad 
        de crear tokens similares con mayor probabilidad.

    Returns:
    --------

    token_hash: str
        Token aleatoria generado en formato string basado en SHA256
    """
    # Creamos el string a partir de la union de los digitos y las letras
    caracteres = string.ascii_letters + string.digits

    # Iniciamos el bucle hasta que se genere un token valido
    while True:

        # Generamos un token aleatorio
        token_raw = ''.join(random.choice(caracteres) for _ in range(length))

        # Añadimos una marca de tiempo para mayor unicidad
        token_timestamp = token_raw + str(time.time())

        # Utilizamos un hash SHA-256 para el token
        token_hash = hashlib.sha256(token_timestamp.encode()).hexdigest()

        # Verificamos si el token ya existe
        if token_hash not in lst_existing_token:

            # En caso de que no exista lo retornamos y finaliza el bucle
            return token_hash

def dataframe_from_tbl_excel(
        path_xls: str,
        name_tbl: str
    ) -> pd.DataFrame:
    """
    Metodo para obtener un Dataframe de la clase Pandas 
    a partir de una ficha excel, esta funcion buscara en 
    toda el libro Excel la tabla que se indique. 
    Es importante saber que si no existe la tabla esta 
    funcion generara un error. Las dimensiones del 
    Dataframe que se retorna sera la misma que contenga 
    la tabla a buscar en libro Excel.

    Args:
    -----
    path_xls: str
        Ruta absoluta del libro Excel sobre el que se 
        quiere obtener la data
    
    name_tbl: str
        Nombre de la tabla que se quiere extraer en forma 
        de Dataframe del libro Excel

    Returns:
    --------
    Dataframe:
        Dataframe con los datos encontrados en el libro 
        Excel segun la tabla indicada.
    """
    # Cargamos el archivo Excel
    wb = openpyxl.load_workbook(path_xls, data_only=True)
    
    # Buscamos la tabla por su nombre en cada hoja del libro 
    # instanciando la variable que almacenara la tabla
    tabla = None
    for wb_sheet in wb.sheetnames:
        sheet = wb[wb_sheet]

        # Ahora por cada tabla existente en la hoja actual
        for table in sheet.tables.values():
            if table.name == name_tbl:
                tabla = table
                break
        if tabla:
            break

    # En caso de que no exista ninguna tabla con el nombre indicado
    if not tabla:
        raise ValueError(
            f"No se encontró una tabla con el nombre '{name_tbl}'"
            )

    # Calculamos las dimensiones de la tabla
    min_col, min_row,\
        max_col, max_row = openpyxl.utils.range_boundaries(
            tabla.ref
            )

    # Leemos los datos de la tabla para luego instanciar el DataFrame
    data = sheet.iter_rows(
        min_row = min_row,
        max_row = max_row,
        min_col = min_col,
        max_col = max_col,
        values_only = True
        )
    return pd.DataFrame(data, columns=next(data))